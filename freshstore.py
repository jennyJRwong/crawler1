# coding=utf-8
from openpyxl import Workbook
from openpyxl import load_workbook
# -*- coding: utf-8 -*-
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException
import re
import pandas as pd
import time
from bs4 import BeautifulSoup


f = open('information.txt','a')
class getfreahstoreAdress:
    def __init__(self,cookie):
        self.path = r'/usr/local/bin/chromedriver'
        # self._cookie = self._format_cookies(cookie)
        self.url = 'http://api.map.baidu.com/lbsapi/getpoint/index.html'
    # 初始化历览器
    def _init_browser(self):
        """
            初始化游览器
        """
        chrome_options = Options()
        # chrome_options.add_argument('--headless')
        # chrome_options.add_argument('--disable-gpu')
        # 禁止加载图片
        prefs = {
            'profile.default_content_setting_values': {
                'images': 2
            }
        }
        chrome_options.add_experimental_option('prefs', prefs)
        browser = webdriver.Chrome(chrome_options=chrome_options, executable_path=self.path)
        # for name, value in self._cookie.items():
        #     browser.add_cookie({'name': name, 'value': value})
        #     browser.refresh()
        return browser

    def _format_cookies(self, cookies):
        """
        获取cookies;;;
        :param cookies:
        :return:
        """
        cookies = {cookie.split('=')[0]: cookie.split('=')[1] for cookie in cookies.replace('', '').split(';')}
        return cookies

    def coordinate(self):
        # 创建浏览器驱动对象
        driver = self._init_browser()
        driver.get(self.url)
        # 显式等待，设置timeout
        wait = WebDriverWait(driver, 3)  # 等待的最大时间
        # 判断输入框是否加载
        input = wait.until(
            EC.presence_of_element_located(
                (By.CSS_SELECTOR, '#localvalue')))
        # 判断搜索按钮是否加载
        submit = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#localsearch')))
        # 输入搜索词，点击搜索按钮

        # 有时候我们希望读取到公式计算出来的结果，可以使用load_workbook()中的data_only属性
        wb = load_workbook(u'./shop.xlsx', data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows():
            rows.append(row)
        print(u"行高：", ws.max_row)
        print(u"列宽：", ws.max_column)
        for i in range(1, ws.max_row):  # row
            # print  rows[i][0], rows[i][0].value, type(rows[i][0].value)
            print(rows[i][1], rows[i][1].value)
            # coordinate(rows[i][0].value)
            input.clear()
            input.send_keys(rows[i][1].value)
            submit.click()
            time.sleep(1)
            try:
                # 等待坐标
                wait.until(
                    EC.presence_of_element_located(
                        (By.CSS_SELECTOR, '#no_0')))
            except TimeoutException:
                print('百度地图查不到地址')
                continue
            # 获取网页文本，提取经纬度
            source = driver.page_source
            soup = BeautifulSoup(source, 'lxml')  #
            i = 0
            for li in soup.select('ul.local_s > li'):
                print(li.get_text())
                f.write(li.get_text())
                f.write('\n')
                i += 1
                if i > 0:
                    break

        # 关闭浏览器驱动
        driver.close()

    def main(self):
        self.coordinate()

if __name__ == '__main__':
    Cookie = 'BAIDUID=2504F198915257ECB8C910053D7987A8:FG=1; ' \
             'BDUSS=0I3aUhIZnRYdGNVcG5pWXBDZzc0Q0F5YVVHdHVVTGY0b1dsZHhPZFNqeGM2UkplRVFBQUFBJCQAAAAAAAAAAAEAAACkiNRKvfDT47rN1u3W7QAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFxc611cXOtdS; pgv_pvi=9871870976; ' \
             'PSTM=1583631843; BIDUPSID=871BAF6EF3620EF44A5D98D4D7295956; ' \
             'H_PS_PSSID=30975_1465_21078_30907_30823_31086_26350; ' \
             'MCITY=-%3A'
    freshStore = getfreahstoreAdress(Cookie)
    freshStore.main()




